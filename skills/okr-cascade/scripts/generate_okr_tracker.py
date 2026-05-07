#!/usr/bin/env python3
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def parse_notion_structure(content):
    """Parse Notion markdown into structured hierarchy."""
    lines = content.split('\n')
    structure = {
        'company_objectives': [],
        'portfolios': {},
        'teams': []
    }
    
    current_section = None
    current_portfolio = None
    current_team = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if line.startswith('# Company Objectives'):
            current_section = 'company_objectives'
            current_portfolio = None
            current_team = None
        elif line.startswith('# Portfolio:'):
            current_section = 'portfolio'
            current_portfolio = line.replace('# Portfolio:', '').strip()
            structure['portfolios'][current_portfolio] = {'name': current_portfolio, 'krs': []}
            current_team = None
        elif line.startswith('## Team'):
            current_section = 'team'
            match = re.search(r'Team (\d+):\s*(.+)', line)
            if match:
                team_num = int(match.group(1))
                team_name = match.group(2)
                current_team = {'number': team_num, 'name': team_name, 'objective': '', 'krs': [], 'owner': ''}
                structure['teams'].append(current_team)
        elif current_section == 'company_objectives' and line.startswith('-'):
            objective = line.lstrip('- ').strip()
            if objective:
                structure['company_objectives'].append(objective)
        elif current_section == 'portfolio' and current_portfolio:
            if line.startswith('-') or line.startswith('KRs:'):
                kr = line.lstrip('- KRs: ').strip()
                if kr:
                    structure['portfolios'][current_portfolio]['krs'].append(kr)
        elif current_section == 'team' and current_team:
            if line.startswith('- Objective:'):
                current_team['objective'] = line.replace('- Objective:', '').strip()
            elif line.startswith('- KRs:'):
                krs_str = line.replace('- KRs:', '').strip()
                current_team['krs'] = [kr.strip() for kr in krs_str.split(',')]
            elif line.startswith('- Owner:'):
                current_team['owner'] = line.replace('- Owner:', '').strip()
    
    return structure

def create_excel_tracker(structure, quarter):
    """Generate pre-populated Excel tracker."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # ===== PORTFOLIO VIEW =====
    portfolio = wb.create_sheet("Portfolio View", 0)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        portfolio.column_dimensions[col].width = 18
    
    row = 1
    portfolio[f'A{row}'] = "PORTFOLIO VIEW"
    portfolio[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    portfolio[f'A{row}'] = f"Quarter: {quarter}"
    row += 1
    portfolio[f'A{row}'] = "Last Updated:"
    portfolio[f'B{row}'] = "=TODAY()"
    row += 2
    
    # Company Objectives
    portfolio[f'A{row}'] = "COMPANY OBJECTIVES"
    portfolio[f'A{row}'].font = subheader_font
    portfolio[f'A{row}'].fill = subheader_fill
    row += 1
    
    headers = ["#", "Objective", "Owner"]
    for col, header in enumerate(headers, 1):
        cell = portfolio.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    for i, obj in enumerate(structure['company_objectives'], 1):
        portfolio[f'A{row}'] = i
        portfolio[f'B{row}'] = obj
        row += 1
    
    row += 2
    
    # Portfolio KRs
    portfolio[f'A{row}'] = "PORTFOLIO KEY RESULTS"
    portfolio[f'A{row}'].font = subheader_font
    portfolio[f'A{row}'].fill = subheader_fill
    row += 1
    
    headers = ["#", "Portfolio KR", "Owner", "Baseline", "Target", "Current", "Score"]
    for col, header in enumerate(headers, 1):
        cell = portfolio.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    for portfolio_name, portfolio_data in structure['portfolios'].items():
        for i, kr in enumerate(portfolio_data['krs'], 1):
            portfolio[f'A{row}'] = len(structure['company_objectives']) + i
            portfolio[f'B{row}'] = kr
            portfolio[f'C{row}'] = ""
            portfolio[f'D{row}'] = 0
            portfolio[f'E{row}'] = 0
            portfolio[f'F{row}'] = 0
            portfolio[f'G{row}'] = f'=(F{row}-D{row})/(E{row}-D{row})'
            for col in range(1, 8):
                portfolio.cell(row=row, column=col).border = thin_border
            row += 1
    
    portfolio.freeze_panes = "A4"
    
    # ===== TEAM TABS =====
    for team in structure['teams']:
        team_num = team['number']
        team_name = team['name']
        team_sheet = wb.create_sheet(f"Team {team_num}", team_num)
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            team_sheet.column_dimensions[col].width = 15
        
        row = 1
        team_sheet[f'A{row}'] = f"TEAM {team_num}: {team_name}"
        team_sheet[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        team_sheet[f'A{row}'] = "Objective:"
        team_sheet[f'B{row}'] = team['objective']
        row += 1
        team_sheet[f'A{row}'] = "Owner:"
        team_sheet[f'B{row}'] = team['owner']
        row += 1
        team_sheet[f'A{row}'] = "Quarter:"
        team_sheet[f'B{row}'] = quarter
        row += 2
        
        team_sheet[f'A{row}'] = "KEY RESULTS"
        team_sheet[f'A{row}'].font = subheader_font
        team_sheet[f'A{row}'].fill = subheader_fill
        row += 1
        
        headers = ["#", "Key Result", "Baseline", "Target", "M1", "M2", "M3", "Score", "Confidence", "Owner"]
        for col, header in enumerate(headers, 1):
            cell = team_sheet.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        row += 1
        
        for i, kr in enumerate(team['krs'][:5], 1):
            team_sheet[f'A{row}'] = i
            team_sheet[f'B{row}'] = kr
            team_sheet[f'C{row}'] = 0
            team_sheet[f'D{row}'] = 0
            team_sheet[f'H{row}'] = f'=IF(G{row}<>"",G{row},IF(F{row}<>"",F{row},IF(E{row}<>"",E{row},"")))'
            
            for col in range(1, 11):
                cell = team_sheet.cell(row=row, column=col)
                cell.border = thin_border
                if col in [3, 4, 5, 6, 7, 8]:
                    cell.number_format = '0.0'
                if i % 2 == 0:
                    cell.fill = alt_row_fill
            row += 1
        
        team_sheet.freeze_panes = "A4"
    
    # ===== ALIGNMENT CHECK =====
    alignment = wb.create_sheet("Alignment Check", 1)
    alignment.column_dimensions['A'].width = 15
    alignment.column_dimensions['B'].width = 30
    
    row = 1
    alignment[f'A{row}'] = "ALIGNMENT CHECK"
    alignment[f'A{row}'].font = Font(bold=True, size=14)
    row += 2
    
    headers = ["Team", "Key Result", "Linked Portfolio KR", "Alignment Strength"]
    for col, header in enumerate(headers, 1):
        cell = alignment.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    
    for team in structure['teams']:
        for kr in team['krs'][:3]:
            alignment[f'A{row}'] = team['name']
            alignment[f'B{row}'] = kr
            alignment[f'D{row}'] = ""
            row += 1
    
    alignment.freeze_panes = "A4"
    
    return wb

def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_okr_tracker.py <quarter> [notion_content]")
        sys.exit(1)
    
    quarter = sys.argv[1]
    content = sys.argv[2] if len(sys.argv) > 2 else sys.stdin.read()
    
    structure = parse_notion_structure(content)
    wb = create_excel_tracker(structure, quarter)
    
    filename = f"OKR_Tracker_{quarter.replace(" ", "_")}.xlsx"
    wb.save(filename)
    print(f"✓ Generated: {filename}")

if __name__ == "__main__":
    main()
